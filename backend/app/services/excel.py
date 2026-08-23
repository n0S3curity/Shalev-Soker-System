"""Server-side Excel generation.

Two workbooks, one sheet per municipality:

* full data   - every survey field, with the signature and photos embedded as
                real pictures rather than pasted base64 text
* calculation - the "תחשיב" layout: one breakdown row per container, a summary
                row per business, and live Excel formulas the operator fills in
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any, Iterable, Optional

from bson import ObjectId
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from .. import db
from ..constants import freq_to_monthly

log = logging.getLogger("excel")

HEADER_FILL = PatternFill("solid", fgColor="27500A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUMMARY_FILL = PatternFill("solid", fgColor="EAF3DE")
SUMMARY_FONT = Font(bold=True, color="173404")
THIN = Side(style="thin", color="C9D3BE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

IMAGE_THUMB = (150, 110)
IMAGE_ROW_HEIGHT = 88
MAX_EMBEDDED_IMAGES = 3
MAX_DOC_SLOTS = 4


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join("_" if ch in ':\\/?*[]' else ch for ch in (name or "ללא עיר"))
    return cleaned[:31] or "גיליון"


def _yes_no(value: str) -> str:
    return {"yes": "כן", "no": "לא"}.get(value or "", value or "")


def _fmt_date(value: Optional[datetime]) -> str:
    if not value:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.strftime("%d/%m/%Y %H:%M")


def _style_header(worksheet, columns: int) -> None:
    for index in range(1, columns + 1):
        cell = worksheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 34
    worksheet.freeze_panes = "A2"


def _group_by_city(records: Iterable[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        grouped.setdefault(record.get("city") or "ללא עיר", []).append(record)
    return grouped


async def _load_thumbnail(file_id: Any) -> Optional[bytes]:
    """Fetch an upload from GridFS and shrink it so the workbook stays usable."""
    if not isinstance(file_id, ObjectId):
        return None
    try:
        stream = await db.gridfs().open_download_stream(file_id)
        raw = await stream.read()
        image = PILImage.open(io.BytesIO(raw))
        image = image.convert("RGB")
        image.thumbnail(IMAGE_THUMB, PILImage.LANCZOS)
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        return buffer.getvalue()
    except Exception:
        log.warning("could not embed file %s", file_id, exc_info=True)
        return None


async def _doc_names(ids: list[Any]) -> list[str]:
    if not ids:
        return []
    records = await db.get_db()["uploads.files"].find(
        {"_id": {"$in": [i for i in ids if isinstance(i, ObjectId)]}}
    ).to_list(50)
    by_id = {r["_id"]: r.get("filename", "") for r in records}
    return [by_id.get(i, "") for i in ids if isinstance(i, ObjectId)]


def _container_detail(containers: list[dict[str, Any]]) -> str:
    lines = []
    for entry in containers or []:
        volume = f" {entry.get('vol')}ל׳" if entry.get("vol") else ""
        freq = entry.get("freq") or "—"
        if freq == "אחר" and entry.get("freqOther"):
            freq = entry["freqOther"]
        lines.append(
            f"{entry.get('ctype', '')}{volume} ×{entry.get('qty') or '?'} | "
            f"תדירות:{freq} | בעלות:{entry.get('ownership') or '—'} | "
            f"שימוש:{entry.get('usage') or '—'}"
        )
    return "\n".join(lines)


# =========================================================================
#  Workbook A - full data
# =========================================================================
FULL_HEADERS = [
    "שם העסק", "מספר ח.פ/ע.מ/ת.ז", "שם הנציג", "תפקיד", "טלפון נציג", "טלפון עסק",
    "כתובת", "עיר", "מהות העסק", "מרכיב התאמה", "מטבח פעיל", "חצר", "גודל חצר",
    "כלי אצירה (פירוט)", "בעלות", "שימוש", "פסולת מעורבת", "פסולת קרטון",
    "נמסרה הצהרה", "הוחזרה הצהרה", "מספר עובדים", "הערות",
    "ממלא הסקר", "נוצר בתאריך", "עודכן בתאריך",
    "חתימה", "תמונה 1", "תמונה 2", "תמונה 3",
    "מסמך 1", "מסמך 2", "מסמך 3", "מסמך 4",
]

FULL_WIDTHS = [
    22, 18, 16, 14, 15, 15, 28, 12, 18, 13, 10, 8, 10,
    52, 14, 14, 13, 13, 12, 13, 12, 34,
    22, 18, 18,
    24, 24, 24, 24,
    26, 26, 26, 26,
]


async def build_full_workbook(records: list[dict[str, Any]], embed_images: bool = True) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    grouped = _group_by_city(records)
    if not grouped:
        grouped = {"ללא נתונים": []}

    for city, city_records in sorted(grouped.items()):
        worksheet = workbook.create_sheet(_safe_sheet_name(city))
        worksheet.sheet_view.rightToLeft = True
        worksheet.append(FULL_HEADERS)
        _style_header(worksheet, len(FULL_HEADERS))

        for index, width in enumerate(FULL_WIDTHS, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width

        anchored: list[XLImage] = []

        for row_index, record in enumerate(city_records, start=2):
            containers = record.get("containers") or []
            docs = await _doc_names(record.get("doc_ids", []))

            values = [
                record.get("biz_name", ""),
                record.get("biznum", ""),
                record.get("rep_name", ""),
                record.get("role", ""),
                record.get("rep_phone", ""),
                record.get("biz_phone", ""),
                record.get("address", ""),
                record.get("city", ""),
                record.get("biz_type", ""),
                record.get("sector", ""),
                _yes_no(record.get("kitchen", "")),
                _yes_no(record.get("yard", "")),
                record.get("yard_size", ""),
                _container_detail(containers),
                ", ".join(sorted({c.get("ownership", "") for c in containers if c.get("ownership")})),
                ", ".join(sorted({c.get("usage", "") for c in containers if c.get("usage")})),
                record.get("wet", ""),
                record.get("cardboard", ""),
                _yes_no(record.get("decl_given", "")),
                _yes_no(record.get("decl_ret", "")),
                record.get("emp_count", ""),
                record.get("notes", ""),
                record.get("owner_name") or record.get("owner_email", ""),
                _fmt_date(record.get("created_at")),
                _fmt_date(record.get("updated_at")),
                "",  # signature - filled with a picture below
            ]
            values.extend([""] * MAX_EMBEDDED_IMAGES)
            values.extend((docs + [""] * MAX_DOC_SLOTS)[:MAX_DOC_SLOTS])

            worksheet.append(values)

            for column in range(1, len(FULL_HEADERS) + 1):
                cell = worksheet.cell(row=row_index, column=column)
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="right")

            if not embed_images:
                continue

            worksheet.row_dimensions[row_index].height = IMAGE_ROW_HEIGHT
            media: list[tuple[int, Any]] = []
            if record.get("signature_id"):
                media.append((26, record["signature_id"]))
            for offset, image_id in enumerate(record.get("image_ids", [])[:MAX_EMBEDDED_IMAGES]):
                media.append((27 + offset, image_id))

            for column, file_id in media:
                thumbnail = await _load_thumbnail(file_id)
                if not thumbnail:
                    worksheet.cell(row=row_index, column=column, value="(הקובץ לא נטען)")
                    continue
                picture = XLImage(io.BytesIO(thumbnail))
                picture.anchor = f"{get_column_letter(column)}{row_index}"
                worksheet.add_image(picture)
                anchored.append(picture)

        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(FULL_HEADERS))}1"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# =========================================================================
#  Workbook B - calculation (תחשיב)
# =========================================================================
CALC_HEADERS = [
    "מס׳ סידורי",              # A
    "שם העסק",                 # B
    "מספר העסק",               # C
    "מהות העסק",               # D
    "שם הנציג",                # E
    "טלפון הנציג",             # F
    "טלפון בעסק",              # G
    "כתובת",                   # H
    "מספר עובדים",             # I
    'ק"ג לעובד',               # J - operator fills
    "מרכיב התאמה",             # K
    "מרכיב עידוד מחזור",       # L
    'פסולת בסיסית ק"ג/חודש',   # M
    'עלות ק"ג',                # N - operator fills
    "זיכוי פסולת בסיסית/חודש", # O
    "כמות כלי אצירה",          # P
    "סוג כלי אצירה",           # Q
    "עלות לפינוי",             # R - operator fills
    "מקדם נפח",                # S - operator fills
    "תדירות פינוי/חודש",       # T
    "עלות פסולת כוללת/חודש",   # U
    'סה"כ תשלום לחודש',        # V
]

CALC_WIDTHS = [7, 24, 16, 18, 16, 15, 15, 26, 9, 9, 11, 12, 15, 10, 15, 10, 22, 10, 10, 13, 15, 15]


def _sector_modifier(record: dict[str, Any]) -> Any:
    sector = record.get("sector")
    if sector == "תעשייה":
        return 1.5
    if sector == "מסחר":
        return 2
    return ""


def _recycling_modifier(record: dict[str, Any]) -> Any:
    cardboard = record.get("cardboard")
    if cardboard == "לא מפנה":
        return 0.8
    if cardboard in ("רשות", "עצמי"):
        return 1
    return ""


def build_calc_workbook(records: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)

    grouped = _group_by_city(records)
    if not grouped:
        grouped = {"ללא נתונים": []}

    for city, city_records in sorted(grouped.items()):
        worksheet = workbook.create_sheet(_safe_sheet_name(city))
        worksheet.sheet_view.rightToLeft = True
        worksheet.append(CALC_HEADERS)
        _style_header(worksheet, len(CALC_HEADERS))

        for index, width in enumerate(CALC_WIDTHS, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width

        excel_row = 2
        serial = 1

        for record in city_records:
            containers = record.get("containers") or []
            entries: list[Optional[dict[str, Any]]] = list(containers) if containers else [None]

            first_row = excel_row
            last_row = excel_row + len(entries) - 1
            summary_row = last_row + 1

            employees = record.get("emp_count") or ""
            try:
                employees = float(employees) if employees else ""
            except ValueError:
                employees = ""

            for position, entry in enumerate(entries):
                row_number = excel_row
                row: list[Any] = [""] * 22

                row[0] = serial if position == 0 else ""
                row[1] = record.get("biz_name", "")           # repeated for filtering
                row[2] = record.get("biznum", "") if position == 0 else ""
                row[3] = record.get("biz_type", "") if position == 0 else ""
                row[4] = record.get("rep_name", "") if position == 0 else ""
                row[5] = record.get("rep_phone", "") if position == 0 else ""
                row[6] = record.get("biz_phone", "") if position == 0 else ""
                row[7] = record.get("address", "") if position == 0 else ""

                row[8] = employees if position == 0 else ""
                row[9] = ""
                row[10] = _sector_modifier(record) if position == 0 else ""
                row[11] = _recycling_modifier(record) if position == 0 else ""
                row[12] = (
                    f"=I{row_number}*J{row_number}*K{row_number}*L{row_number}"
                    if position == 0 else ""
                )
                row[13] = ""
                row[14] = f"=M{row_number}*N{row_number}"

                if entry:
                    row[15] = entry.get("qty") or ""
                    volume = f" {entry.get('vol')}" if entry.get("vol") else ""
                    row[16] = f"{entry.get('ctype', '')}{volume}".strip()
                    monthly = freq_to_monthly(entry.get("freq"))
                    row[19] = monthly if monthly != "" else ""

                row[20] = f"=T{row_number}*S{row_number}*R{row_number}*P{row_number}"

                worksheet.append(row)
                for column in range(1, 23):
                    worksheet.cell(row=row_number, column=column).border = BORDER
                excel_row += 1

            summary: list[Any] = [""] * 22
            summary[1] = record.get("biz_name", "")
            summary[14] = f"=SUM(O{first_row}:O{last_row})"
            summary[15] = f"=SUM(P{first_row}:P{last_row})"
            summary[16] = 'סה"כ כלי אצירה לעסק'
            summary[19] = f"=SUM(T{first_row}:T{last_row})"
            summary[20] = f"=SUM(U{first_row}:U{last_row})"
            summary[21] = f"=U{summary_row}-O{first_row}"

            worksheet.append(summary)
            for column in range(1, 23):
                cell = worksheet.cell(row=summary_row, column=column)
                cell.fill = SUMMARY_FILL
                cell.font = SUMMARY_FONT
                cell.border = BORDER
            excel_row += 1

            worksheet.append([""] * 22)  # separator between businesses
            excel_row += 1
            serial += 1

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# =========================================================================
#  Query helper shared by the export endpoints and the daily mail job
# =========================================================================
async def fetch_records(
    city: Optional[str] = None,
    since: Optional[datetime] = None,
    limit: int = 20000,
) -> list[dict[str, Any]]:
    query: dict[str, Any] = {"deleted": False}
    if city:
        query["city"] = city
    if since:
        query["updated_at"] = {"$gte": since}
    cursor = db.surveys().find(query).sort([("city", 1), ("biz_name", 1)])
    return await cursor.to_list(limit)
